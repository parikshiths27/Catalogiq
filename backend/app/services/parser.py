import uuid
import os
import io
import csv
import json
import tempfile
import logging
import xml.etree.ElementTree as ET
from abc import ABC, abstractmethod
from typing import Optional, Dict, Any, List

logger = logging.getLogger(__name__)


class DocumentParser(ABC):
    @abstractmethod
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        """
        Parses document binary content and returns a structured intermediate representation.
        """
        pass


class DoclingParser(DocumentParser):
    def __init__(self):
        # Force docling import to fail clearly at runtime if unavailable
        try:
            import docling
            from docling.document_converter import DocumentConverter
            self._converter_class = DocumentConverter
            self.version = docling.__version__
        except ImportError as e:
            logger.error("Docling library not available at runtime.")
            raise ImportError(
                "Docling library is not installed or available at runtime. "
                "Ensure 'docling' is listed in requirements and installed."
            ) from e

    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        ext = os.path.splitext(filename.lower())[1] if filename else ".pdf"
        if ext not in (".pdf", ".docx"):
            ext = ".pdf"

        # Write binary stream to a temporary local file with matching extension for Docling
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name

        try:
            converter = self._converter_class()
            result = converter.convert(tmp_path)
            doc = result.document

            pages: List[Dict[str, Any]] = []
            if hasattr(doc, "num_pages"):
                page_count = doc.num_pages() if callable(doc.num_pages) else doc.num_pages
            else:
                page_count = 1
            
            # Pre-initialize page mappings
            for page_idx in range(1, max(page_count, 1) + 1):
                pages.append({
                    "page_number": page_idx,
                    "text": "",
                    "tables": [],
                    "images": []
                })

            # Process layout elements and associate them with correct page boundaries
            for element, level in doc.iterate_items():
                page_no = 1
                if hasattr(element, "prov") and element.prov:
                    page_no = element.prov[0].page_no if hasattr(element.prov[0], "page_no") else 1
 
                # Ensure page bounds in pages array
                if page_no > len(pages):
                    while len(pages) < page_no:
                        pages.append({
                            "page_number": len(pages) + 1,
                            "text": "",
                            "tables": [],
                            "images": []
                        })
                
                page_data = pages[page_no - 1]
                class_name = element.__class__.__name__
                
                if "Table" in class_name:
                    headers = []
                    rows = []
                    try:
                        df = element.export_to_dataframe(doc)
                        headers = [str(col) for col in df.columns]
                        rows = [[str(val) for val in row] for row in df.values.tolist()]
                    except Exception as df_err:
                        logger.warning(f"export_to_dataframe failed, falling back to manual cell parsing: {df_err}")
                        grid = getattr(element, "data", None)
                        if grid and hasattr(grid, "table_cells"):
                            from collections import defaultdict
                            row_cells = defaultdict(list)
                            for cell in grid.table_cells:
                                row_cells[cell.start_row_offset_idx].append(cell)
                            for r_idx in sorted(row_cells.keys()):
                                sorted_cells = sorted(row_cells[r_idx], key=lambda c: c.start_col_offset_idx)
                                row_vals = [c.text for c in sorted_cells]
                                if r_idx == 0:
                                    headers = row_vals
                                else:
                                    rows.append(row_vals)
                    page_data["tables"].append({
                        "headers": headers,
                        "rows": rows
                    })
                elif "Picture" in class_name or "Image" in class_name:
                    page_data["images"].append({
                        "image_id": str(uuid.uuid4()),
                        "page_number": page_no,
                        "label": getattr(element, "label", "image")
                    })
                else:
                    text_val = getattr(element, "text", "")
                    if text_val:
                        page_data["text"] += (text_val + "\n")

            title = getattr(doc, "title", None) or (filename or "Technical Specification Document")
            return {
                "pages": pages,
                "metadata": {
                    "page_count": len(pages),
                    "title": title
                }
            }

        except Exception as e:
            import traceback
            logger.error(f"Error during Docling parsing: {e}\n{traceback.format_exc()}")
            raise e
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass


class ExcelParser(DocumentParser):
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        pages: List[Dict[str, Any]] = []

        for idx, sheet_name in enumerate(wb.sheetnames, start=1):
            sheet = wb[sheet_name]
            all_rows = []
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_vals):
                    all_rows.append(row_vals)

            headers = all_rows[0] if all_rows else []
            data_rows = all_rows[1:] if len(all_rows) > 1 else []

            text_lines = [f"Sheet: {sheet_name}"]
            for r in all_rows:
                text_lines.append(" | ".join(r))

            tables = []
            if headers:
                tables.append({"headers": headers, "rows": data_rows})

            pages.append({
                "page_number": idx,
                "text": "\n".join(text_lines) + "\n",
                "tables": tables,
                "images": []
            })

        if not pages:
            pages = [{"page_number": 1, "text": "", "tables": [], "images": []}]

        return {
            "pages": pages,
            "metadata": {
                "page_count": len(pages),
                "title": filename or "Excel Spreadsheet Catalog"
            }
        }


class CSVParser(DocumentParser):
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text = file_content.decode("latin-1", errors="replace")

        reader = csv.reader(io.StringIO(text))
        rows = [row for row in reader if any(cell.strip() for cell in row)]

        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        tables = []
        if headers:
            tables.append({"headers": headers, "rows": data_rows})

        return {
            "pages": [{
                "page_number": 1,
                "text": text,
                "tables": tables,
                "images": []
            }],
            "metadata": {
                "page_count": 1,
                "title": filename or "CSV Product Catalog Data"
            }
        }


class TextParser(DocumentParser):
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text = file_content.decode("latin-1", errors="replace")

        lines = text.splitlines()
        page_size = 60
        pages: List[Dict[str, Any]] = []

        if not lines:
            pages.append({"page_number": 1, "text": "", "tables": [], "images": []})
        else:
            for page_idx in range(0, len(lines), page_size):
                chunk = lines[page_idx:page_idx + page_size]
                pages.append({
                    "page_number": (page_idx // page_size) + 1,
                    "text": "\n".join(chunk) + "\n",
                    "tables": [],
                    "images": []
                })

        return {
            "pages": pages,
            "metadata": {
                "page_count": len(pages),
                "title": filename or "Technical Specification Document"
            }
        }


class JSONParser(DocumentParser):
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text = file_content.decode("latin-1", errors="replace")

        data = json.loads(text)
        tables: List[Dict[str, Any]] = []

        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            # Extract headers from dictionary keys across items
            keys = list({k for d in data if isinstance(d, dict) for k in d.keys()})
            rows = [[str(d.get(k, "")) for k in keys] for d in data if isinstance(d, dict)]
            tables.append({"headers": keys, "rows": rows})
        elif isinstance(data, dict):
            rows = [[str(k), str(v)] for k, v in data.items() if not isinstance(v, (dict, list))]
            if rows:
                tables.append({"headers": ["Attribute", "Value"], "rows": rows})

        return {
            "pages": [{
                "page_number": 1,
                "text": json.dumps(data, indent=2),
                "tables": tables,
                "images": []
            }],
            "metadata": {
                "page_count": 1,
                "title": filename or "JSON Catalog Export"
            }
        }


class XMLParser(DocumentParser):
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text = file_content.decode("latin-1", errors="replace")

        root = ET.fromstring(text)
        tables: List[Dict[str, Any]] = []
        text_parts: List[str] = [f"Root Tag: {root.tag}"]

        # Process repeated child records (e.g. <product>, <item>, <spec>)
        child_records = list(root)
        if child_records:
            sample_keys = list({sub.tag for child in child_records for sub in child})
            if sample_keys:
                rows = []
                for child in child_records:
                    row_vals = [child.findtext(tag, default="") for tag in sample_keys]
                    rows.append(row_vals)
                tables.append({"headers": sample_keys, "rows": rows})

        # Also extract key-value attributes
        if root.attrib:
            attr_rows = [[k, v] for k, v in root.attrib.items()]
            tables.append({"headers": ["Attribute", "Value"], "rows": attr_rows})

        for elem in root.iter():
            if elem.text and elem.text.strip():
                text_parts.append(f"{elem.tag}: {elem.text.strip()}")

        return {
            "pages": [{
                "page_number": 1,
                "text": "\n".join(text_parts),
                "tables": tables,
                "images": []
            }],
            "metadata": {
                "page_count": 1,
                "title": filename or "XML Product Document"
            }
        }


class HTMLParser(DocumentParser):
    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        from bs4 import BeautifulSoup

        try:
            html_text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            html_text = file_content.decode("latin-1", errors="replace")

        soup = BeautifulSoup(html_text, "html.parser")

        # Remove script and style elements
        for element in soup(["script", "style"]):
            element.decompose()

        tables: List[Dict[str, Any]] = []
        for html_table in soup.find_all("table"):
            headers = [th.get_text(strip=True) for th in html_table.find_all("th")]
            rows = []
            for tr in html_table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in tr.find_all("td")]
                if cells:
                    rows.append(cells)

            if not headers and rows:
                headers = rows[0]
                rows = rows[1:]

            if headers or rows:
                tables.append({"headers": headers, "rows": rows})

        clean_text = soup.get_text(separator="\n", strip=True)

        return {
            "pages": [{
                "page_number": 1,
                "text": clean_text,
                "tables": tables,
                "images": []
            }],
            "metadata": {
                "page_count": 1,
                "title": filename or "HTML Specification Document"
            }
        }


class MultiFormatParser(DocumentParser):
    """
    Master multi-format document parser dispatcher.
    Routes incoming files to the appropriate specialized parser based on extension.
    """
    def __init__(self):
        self._docling_parser = None
        self._excel_parser = ExcelParser()
        self._csv_parser = CSVParser()
        self._text_parser = TextParser()
        self._json_parser = JSONParser()
        self._xml_parser = XMLParser()
        self._html_parser = HTMLParser()
        self._mock_parser = MockParser()

    def _get_docling_parser(self) -> DoclingParser:
        if self._docling_parser is None:
            self._docling_parser = DoclingParser()
        return self._docling_parser

    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        if os.getenv("TEST_MOCK_PARSER") == "true":
            return self._mock_parser.parse(file_content, filename=filename)

        ext = os.path.splitext(filename.lower())[1] if filename else ".pdf"

        if ext in (".pdf", ".docx"):
            return self._get_docling_parser().parse(file_content, filename=filename)
        elif ext == ".xlsx":
            return self._excel_parser.parse(file_content, filename=filename)
        elif ext == ".csv":
            return self._csv_parser.parse(file_content, filename=filename)
        elif ext in (".txt", ".md"):
            return self._text_parser.parse(file_content, filename=filename)
        elif ext == ".json":
            return self._json_parser.parse(file_content, filename=filename)
        elif ext == ".xml":
            return self._xml_parser.parse(file_content, filename=filename)
        elif ext in (".html", ".htm"):
            return self._html_parser.parse(file_content, filename=filename)
        else:
            # Fallback to TextParser for unknown format
            return self._text_parser.parse(file_content, filename=filename)


class MockParser(DocumentParser):
    def __init__(self):
        self.version = "1.0.0"

    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        """
        Mock implementation explicitly injected for tests.
        Supports both PDF and non-PDF mock data generation.
        """
        ext = os.path.splitext(filename.lower())[1] if filename else ".pdf"

        # Format-specific test mock handling
        if ext == ".docx":
            return {
                "pages": [{
                    "page_number": 1,
                    "text": "DOCX Spec Sheet\nModel: MX-DOCX\nSKU: DOCX-900\n",
                    "tables": [{"headers": ["Spec", "Value"], "rows": [["Voltage", "230 V"], ["Power", "7.5 kW"]]}],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "DOCX Spec Sheet"}
            }
        elif ext == ".xlsx":
            return {
                "pages": [{
                    "page_number": 1,
                    "text": "Sheet: Products\nSKU | Name | Voltage | Power\n",
                    "tables": [{"headers": ["SKU", "Name", "Voltage", "Power"], "rows": [["XLS-100", "Excel Motor", "400 V", "11 kW"]]}],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "Excel Product Catalog"}
            }
        elif ext == ".csv":
            return {
                "pages": [{
                    "page_number": 1,
                    "text": "SKU,Name,Voltage,Power\nCSV-200,CSV Motor,230V,3.0kW\n",
                    "tables": [{"headers": ["SKU", "Name", "Voltage", "Power"], "rows": [["CSV-200", "CSV Motor", "230V", "3.0kW"]]}],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "CSV Catalog Data"}
            }
        elif ext in (".txt", ".md"):
            return {
                "pages": [{
                    "page_number": 1,
                    "text": "Text Specification\nModel: TXT-300\nSKU: TXT-300-SPEC\nPower: 15 kW\nVoltage: 480 V\n",
                    "tables": [],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "Text Specification"}
            }
        elif ext == ".json":
            return {
                "pages": [{
                    "page_number": 1,
                    "text": '{"sku": "JSON-400", "name": "JSON Motor", "voltage": "230 V"}',
                    "tables": [{"headers": ["sku", "name", "voltage"], "rows": [["JSON-400", "JSON Motor", "230 V"]]}],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "JSON Catalog Export"}
            }
        elif ext == ".xml":
            return {
                "pages": [{
                    "page_number": 1,
                    "text": "<product><sku>XML-500</sku><name>XML Motor</name></product>",
                    "tables": [{"headers": ["sku", "name"], "rows": [["XML-500", "XML Motor"]]}],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "XML Document"}
            }
        elif ext in (".html", ".htm"):
            return {
                "pages": [{
                    "page_number": 1,
                    "text": "HTML Specs\nModel: HTML-600\n",
                    "tables": [{"headers": ["Attribute", "Value"], "rows": [["Power", "22 kW"]]}],
                    "images": []
                }],
                "metadata": {"page_count": 1, "title": "HTML Document"}
            }

        # Default PDF Mock return (validates PDF magic bytes for test fidelity)
        if not file_content.startswith(b"%PDF"):
            raise ValueError("Invalid PDF magic bytes")

        return {
            "pages": [
                {
                    "page_number": 1,
                    "text": "Industrial Motor\nModel: MX-500\nSKU: MX500-230\n",
                    "tables": [],
                    "images": []
                },
                {
                    "page_number": 2,
                    "text": "Specifications\n",
                    "tables": [
                        {
                            "headers": ["Specification", "Value"],
                            "rows": [
                                ["Voltage", "230 V"],
                                ["Power", "5.5 kW"],
                                ["Speed", "1440 RPM"],
                                ["Weight", "32 kg"]
                            ]
                        }
                    ],
                    "images": [
                        {
                            "image_id": "mock-img-123",
                            "page_number": 2,
                            "label": "motor_wiring"
                        }
                    ]
                }
            ],
            "metadata": {
                "page_count": 2,
                "title": "Industrial Motor Specs"
            }
        }
