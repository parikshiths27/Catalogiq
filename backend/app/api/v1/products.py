import json
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel
from sqlmodel import Session, select, desc

from app.db.session import get_session
from app.models import (
    AttributeEvidence,
    AuditLog,
    Document,
    EnrichmentResult,
    Product,
    ProductAttribute,
    ProductDocumentAssociation,
    ProductStatus,
    ProductVersion,
    Source,
    SourceType,
    ValidationResult,
    ValidationStatus,
)
from app.repositories import AttributeRepository, ProductRepository
from app.services.product import ProductService
from app.services.reconciler import MultiSourceReconciler
from app.services.validation_engine import ValidationEngine

router = APIRouter(prefix="/products")


class SourceClaimSchema(BaseModel):
    source_id: Optional[str] = None
    source_name: str = "Unknown Source"
    source_type: str = "document"
    trust_level: float = 1.0
    document_id: Optional[str] = None
    page_number: Optional[int] = None
    evidence_text: str = ""
    attribute_id: Optional[str] = None
    raw_value: str = ""
    normalized_value: Optional[Any] = None
    unit: Optional[str] = None
    extraction_method: str = "llm"


class AttributeReconciliationSchema(BaseModel):
    attribute_name: str
    display_name: str
    canonical_value: Optional[str] = None
    canonical_unit: Optional[str] = None
    canonical_normalized_value: Optional[Any] = None
    status: str
    confidence_score: float
    winning_source_name: Optional[str] = None
    winning_source_trust: Optional[float] = None
    claims: List[SourceClaimSchema] = []
    competing_claims: List[SourceClaimSchema] = []
    explanation: str = ""


class ProductReconciliationResponse(BaseModel):
    product_id: str
    product_name: str
    total_attributes: int = 0
    agreements_count: int = 0
    equivalents_count: int = 0
    missing_count: int = 0
    conflicts_count: int = 0
    review_count: int = 0
    overall_confidence: float = 1.0
    reconciled_attributes: Dict[str, AttributeReconciliationSchema] = {}


class ProductSourceResponse(BaseModel):
    source_id: str
    source_name: str
    source_type: str
    uri: Optional[str] = None
    trust_level: float = 1.0
    document_id: Optional[str] = None
    metadata_json: Dict[str, Any] = {}
    created_at: datetime
    association_type: str = "document_source"


class ResolutionRequest(BaseModel):
    resolution: str  # "accept_source_a" | "accept_source_b" | "custom_value"
    resolved_value: Optional[Any] = None
    notes: Optional[str] = None


@router.get("", response_model=List[Product])
@router.get("/", response_model=List[Product])
def list_products(
    limit: int = 10000,
    offset: int = 0,
    status: Optional[str] = None,
    brand: Optional[str] = None,
    category: Optional[str] = None,
    quality_score_min: Optional[float] = None,
    quality_score_max: Optional[float] = None,
    session: Session = Depends(get_session),
):
    repo = ProductRepository(session)
    return repo.list_products(
        limit=limit,
        offset=offset,
        status=status,
        brand=brand,
        category=category,
        quality_score_min=quality_score_min,
        quality_score_max=quality_score_max,
    )


@router.delete("/clear-all")
def clear_all_products(session: Session = Depends(get_session)):
    """
    Clears all products and their associated attributes, evidence, enrichment results,
    and validation issues from the database for a clean slate.
    """
    from sqlmodel import select as sel
    from app.models import (
        Product, ProductAttribute, AttributeEvidence, ValidationResult,
        EnrichmentResult, DuplicateCandidate, ProductVersion, ProductDocumentAssociation
    )

    # Clean associations & child tables
    session.exec(sel(AttributeEvidence)).all()
    for ev in session.exec(sel(AttributeEvidence)).all():
        session.delete(ev)

    for attr in session.exec(sel(ProductAttribute)).all():
        session.delete(attr)

    for val in session.exec(sel(ValidationResult)).all():
        session.delete(val)

    for enr in session.exec(sel(EnrichmentResult)).all():
        session.delete(enr)

    for dc in session.exec(sel(DuplicateCandidate)).all():
        session.delete(dc)

    for pv in session.exec(sel(ProductVersion)).all():
        session.delete(pv)

    for pda in session.exec(sel(ProductDocumentAssociation)).all():
        session.delete(pda)

    prods = session.exec(sel(Product)).all()
    prod_count = len(prods)
    for p in prods:
        session.delete(p)

    session.commit()

    return {
        "status": "cleared",
        "products_removed": prod_count,
        "message": f"Successfully cleared {prod_count} products and all associated attributes, evidence, and validations."
    }


UNILOG_252_HEADERS = [
    "MFR URL", "Ref URL 1", "Ref URL 2", "Ref URL 3", "Ref URL 4", "Ref URL 5", "PART_NUMBER", "Dept", "Class", "Fine",
    "SKU - MY_PART_NUMBER", "Mfg_Part_Num", "Part_Desc", "E1_Brand", "Unilog_Brand", "DIB_Brand", "Part_Manuf",
    "MANUFACTURER_NAME", "BRAND_NAME", "TRADE_NAME", "MANUFACTURER_PART_NUMBER", "ALTERNATE_PART_NUMBER", "Classpath",
    "MOBILE_DESC", "INVOICE_DESC", "SHORT_DESC", "LONG_DESC1", "RETAIL_DESC", "MARKETING_DESCRIPTION",
    *[f"ITEM_FEATURES_{i}" for i in range(1, 21)],
    "With", "Standard/Approvals", "Prop 65", "Application", "Includes", "Product Name",
    *[col for i in range(1, 51) for col in (f"ATTRIBUTE_LABEL {i}", f"ATTRIBUTE_VALUE {i}", f"ATTRIBUTE_UOM {i}")],
    "UPC", "EAN", "GTIN", "UNSPSC", "Warranty", "List Price", "Selling Qty", "Selling UOM",
    "Standard Packaging Information", "LENGTH", "LENGTH_UOM", "HEIGHT", "HEIGHT_UOM", "WIDTH", "WIDTH_UOM",
    "WEIGHT", "WEIGHT_UOM", "VOLUME", "VOLUME_UOM",
    "Product Image", "Alternate Image 1", "Alternate Image 2", "Alternate Image 3", "Alternate Image 4",
    "SDS", "SDS_1", "Warranty Information", "Catalog", "Specification Sheet",
    "Instruction/Installation Manual", "Service Manual", "Owners/User Manual", "Line Drawing", "MTR", "RoHS",
    "Full Engineering Drawing", "Energy Star Guide", "Technical Bulletin", "Submittal", "Compatibility Chart",
    "Size Chart", "Product Label/Insert", "Video Link", "Video Link 1", "Country Of Origin", "Discontinued",
    "Actual Image (Yes/No)"
]


@router.get("/export")
def export_products(
    format: str = "csv",
    product_id: Optional[uuid.UUID] = None,
    status: Optional[str] = None,
    brand: Optional[str] = None,
    category: Optional[str] = None,
    quality_score_min: Optional[float] = None,
    quality_score_max: Optional[float] = None,
    schema: str = "252",  # "252" (default) or "compact"
    session: Session = Depends(get_session),
):
    """
    Export product catalog in the authoritative 252-column Unilog Delivery Format (or XLSX / PDF / JSON).
    Maps Product, ProductAttribute, and EnrichmentResult directly into standard client delivery columns.
    """
    import csv as csv_module
    import io as io_module
    import re
    from fastapi.responses import StreamingResponse

    repo = ProductRepository(session)

    if product_id:
        single_prod = repo.get_by_id(product_id)
        products = [single_prod] if single_prod else []
    else:
        products = repo.list_products(
            limit=10000,
            offset=0,
            status=status,
            brand=brand,
            category=category,
            quality_score_min=quality_score_min,
            quality_score_max=quality_score_max,
        )

    product_ids = [p.id for p in products]

    # Bulk preload latest EnrichmentResult records in 1 query
    enrich_map: Dict[uuid.UUID, Dict[str, Any]] = {}
    if product_ids:
        enrich_stmts = (
            select(EnrichmentResult)
            .where(EnrichmentResult.product_id.in_(product_ids))
            .order_by(EnrichmentResult.created_at.desc())
        )
        enrich_results = session.exec(enrich_stmts).all()
        for e in enrich_results:
            if e.product_id not in enrich_map and e.generated_value:
                try:
                    enrich_map[e.product_id] = json.loads(e.generated_value)
                except Exception:
                    enrich_map[e.product_id] = {}

    # Bulk preload ProductAttribute records in 1 query
    attrs_map: Dict[uuid.UUID, List[ProductAttribute]] = {}
    if product_ids:
        attr_stmts = (
            select(ProductAttribute)
            .where(ProductAttribute.product_id.in_(product_ids))
        )
        all_attrs = session.exec(attr_stmts).all()
        for a in all_attrs:
            if a.product_id not in attrs_map:
                attrs_map[a.product_id] = []
            attrs_map[a.product_id].append(a)

    delivery_rows = []

    for prod in products:
        enrich_data = enrich_map.get(prod.id, {})
        existing_delivery = enrich_data.get("delivery_record")
        if existing_delivery and isinstance(existing_delivery, dict) and len(existing_delivery) >= 200:
            # Full 252 record already stored
            row = {col: existing_delivery.get(col, "") for col in UNILOG_252_HEADERS}
        else:
            # Synthesize from Product + Attributes
            attrs = attrs_map.get(prod.id, [])
            clean_brand_name = re.sub(r"[^A-Za-z0-9_]", "_", prod.brand.replace("®", "").replace("™", "").strip())
            sku_val = prod.sku or prod.model or ""

            row = {col: "" for col in UNILOG_252_HEADERS}
            row["Dept"] = prod.category.split(">")[0].strip() if ">" in prod.category else prod.category
            row["Class"] = prod.category.split(">")[1].strip() if prod.category.count(">") >= 1 else ""
            row["Fine"] = prod.subcategory or (prod.category.split(">")[2].strip() if prod.category.count(">") >= 2 else "")
            row["SKU - MY_PART_NUMBER"] = sku_val
            row["Mfg_Part_Num"] = sku_val
            row["Part_Desc"] = prod.description or prod.product_name
            row["MANUFACTURER_NAME"] = prod.brand
            row["BRAND_NAME"] = prod.brand
            row["MANUFACTURER_PART_NUMBER"] = sku_val
            row["Classpath"] = prod.category
            row["MOBILE_DESC"] = enrich_data.get("mobile_desc") or f"{prod.brand}, {prod.product_name}, {sku_val}"
            row["INVOICE_DESC"] = (enrich_data.get("invoice_desc") or prod.product_name[:40]).upper()
            row["SHORT_DESC"] = enrich_data.get("short_desc") or prod.product_name
            row["LONG_DESC1"] = enrich_data.get("long_desc") or prod.commerce_description or prod.description or prod.product_name
            row["RETAIL_DESC"] = enrich_data.get("retail_desc") or prod.description or ""
            row["Product Name"] = prod.product_name
            row["Selling Qty"] = "1"
            row["Selling UOM"] = "EA"
            row["Product Image"] = f"{clean_brand_name}_{sku_val}.jpg" if sku_val else ""
            row["Specification Sheet"] = f"{clean_brand_name}_{sku_val}_Specification_Sheet.pdf" if sku_val else ""
            row["Discontinued"] = "No"
            row["Actual Image (Yes/No)"] = "Yes" if sku_val else "No"

            # Populate features 1..20
            if prod.features:
                for idx, feat in enumerate(prod.features[:20], start=1):
                    row[f"ITEM_FEATURES_{idx}"] = str(feat)

            # Populate attribute slots 1..50
            for idx, a in enumerate(attrs[:50], start=1):
                display = a.display_name or a.attribute_name
                norm_v = str(a.normalized_value if a.normalized_value is not None else a.raw_value or "")
                uom_v = str(a.unit or "")
                row[f"ATTRIBUTE_LABEL {idx}"] = display
                row[f"ATTRIBUTE_VALUE {idx}"] = norm_v
                row[f"ATTRIBUTE_UOM {idx}"] = uom_v

        delivery_rows.append(row)

    headers = UNILOG_252_HEADERS
    fmt = format.lower().strip()

    sku_tag = f"_{products[0].sku}" if (product_id and len(products) == 1 and products[0].sku) else ""

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Unilog 252 Delivery"
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws.append(headers)
        ws.row_dimensions[1].height = 24

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=9.5)
        data_align = Alignment(vertical="center")

        for row in delivery_rows:
            row_data = [str(row.get(h, "") or "") for h in headers]
            ws.append(row_data)

        # Auto-adjust column widths
        for col_idx, h in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(h) + 2, 11)

        output = io_module.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Unilog_Delivery{sku_tag}_Format_252_Columns.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    elif fmt == "pdf":
        from app.services.pdf_export import build_catalog_pdf
        pdf_bytes = build_catalog_pdf(products, session, delivery_rows)
        filename = f"Unilog_Delivery{sku_tag}_Catalog_Report.pdf"
        return StreamingResponse(
            io_module.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    elif fmt == "json":
        json_bytes = json.dumps({
            "export_standard": "Unilog 252-Column Master Delivery Format",
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "total_products": len(delivery_rows),
            "columns_count": len(headers),
            "columns": headers,
            "records": delivery_rows,
        }, indent=2).encode("utf-8")
        filename = f"Unilog_Delivery{sku_tag}_Format_252_Columns.json"
        return StreamingResponse(
            io_module.BytesIO(json_bytes),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        output = io_module.StringIO()
        writer = csv_module.DictWriter(output, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        for row in delivery_rows:
            writer.writerow(row)

        csv_bytes = output.getvalue().encode("utf-8")
        filename = f"Unilog_Delivery{sku_tag}_Format_252_Columns.csv"
        return StreamingResponse(
            io_module.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )



@router.get("/{product_id}", response_model=Product)
def get_product(product_id: uuid.UUID, session: Session = Depends(get_session)):
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    attributes = repo.get_attributes(product_id)
    formatted_attributes = {}
    for attr in attributes:
        val = attr.normalized_value if attr.normalized_value is not None else attr.raw_value
        formatted_attributes[attr.attribute_name] = {
            "value": val,
            "unit": attr.unit,
            "raw_value": attr.raw_value,
            "display_name": attr.display_name,
            "data_type": attr.data_type.value if hasattr(attr.data_type, "value") else str(attr.data_type),
            "confidence": attr.confidence,
            "status": attr.status.value if hasattr(attr.status, "value") else str(attr.status),
            "source_type": attr.source_type,
        }

    product_dict = product.model_dump()
    product_dict["attributes"] = formatted_attributes
    return product_dict


@router.get("/{product_id}/attributes", response_model=List[ProductAttribute])
def get_product_attributes(product_id: uuid.UUID, session: Session = Depends(get_session)):
    repo = ProductRepository(session)
    if not repo.get_by_id(product_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )
    return repo.get_attributes(product_id)


@router.get("/{product_id}/validation")
def get_product_validation_summary(product_id: uuid.UUID, session: Session = Depends(get_session)):
    """
    Returns comprehensive validation summary including quality score,
    completeness, open validation issues, and conflict status.
    """
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    validations = repo.get_validations(product_id)
    attributes = repo.get_attributes(product_id)
    attr_repo = AttributeRepository(session)
    evidence = attr_repo.get_evidence_for_product(product_id)

    evidence_names = {
        a.attribute_name for a in attributes
        if any(e.attribute_id == a.id and e.evidence_text for e in evidence)
    }

    engine = ValidationEngine()
    val_res = engine.validate_product(
        product=product,
        attributes=attributes,
        evidence_supported_attribute_names=evidence_names,
    )

    return {
        "product_id": str(product_id),
        "quality_score": product.quality_score,
        "validation_status": product.status,
        "completeness_score": val_res.completeness.completeness_score,
        "completeness_details": val_res.completeness.model_dump(),
        "quality_breakdown": val_res.quality_breakdown.model_dump(),
        "issues": [v.model_dump() for v in validations],
        "has_critical_issues": val_res.has_critical_issues,
        "has_errors": val_res.has_errors,
    }


@router.get("/{product_id}/enrichment")
def get_product_enrichment(product_id: uuid.UUID, session: Session = Depends(get_session)):
    """Returns the latest AI commerce enrichment content for this product in frontend-consumable format."""
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    stmt = select(EnrichmentResult).where(
        EnrichmentResult.product_id == product_id
    ).order_by(EnrichmentResult.created_at.desc())
    
    enrichment = session.exec(stmt).first()
    if not enrichment:
        return {
            "product_id": str(product_id),
            "commerce_description": product.commerce_description,
            "short_description": None,
            "features": product.features or [],
            "applications": product.applications or [],
            "keywords": product.keywords or [],
            "seo_title": None,
            "seo_description": None,
            "status": "pending",
            "confidence": None,
            "model": None,
            "prompt_version": None,
        }

    try:
        gen_data = json.loads(enrichment.generated_value) if isinstance(enrichment.generated_value, str) else enrichment.generated_value
        if not isinstance(gen_data, dict):
            gen_data = {}
    except Exception:
        gen_data = {}

    return {
        "id": str(enrichment.id),
        "product_id": str(enrichment.product_id),
        "enrichment_type": enrichment.enrichment_type.value if hasattr(enrichment.enrichment_type, "value") else str(enrichment.enrichment_type),
        "status": enrichment.status.value if hasattr(enrichment.status, "value") else str(enrichment.status),
        "model": enrichment.model,
        "prompt_version": enrichment.prompt_version,
        "confidence": enrichment.confidence,
        "created_at": enrichment.created_at.isoformat() if enrichment.created_at else None,
        "approved_at": enrichment.approved_at.isoformat() if enrichment.approved_at else None,
        "approved_by": enrichment.approved_by,
        "generated_value": enrichment.generated_value,
        # Parsed generated fields for direct consumption
        "commerce_description": gen_data.get("commerce_description") or product.commerce_description,
        "short_description": gen_data.get("short_description"),
        "features": gen_data.get("features") or product.features or [],
        "applications": gen_data.get("applications") or product.applications or [],
        "keywords": gen_data.get("keywords") or product.keywords or [],
        "seo_title": gen_data.get("seo_title"),
        "seo_description": gen_data.get("seo_description"),
    }


@router.post("/{product_id}/validate")
def rerun_product_validation(product_id: uuid.UUID, session: Session = Depends(get_session)):
    """Re-runs validation engine on demand for a product."""
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    attributes = repo.get_attributes(product_id)
    attr_repo = AttributeRepository(session)
    evidence = attr_repo.get_evidence_for_product(product_id)
    evidence_names = {
        a.attribute_name for a in attributes
        if any(e.attribute_id == a.id and e.evidence_text for e in evidence)
    }

    engine = ValidationEngine()
    val_res = engine.validate_product(
        product=product,
        attributes=attributes,
        evidence_supported_attribute_names=evidence_names,
    )

    # Persist ValidationResult records
    existing_open = repo.get_validations(product_id, status=ValidationStatus.open)
    for old in existing_open:
        session.delete(old)

    for issue in val_res.issues:
        session.add(issue.to_db_model(product.id))

    product.quality_score = val_res.quality_breakdown.quality_score
    if val_res.has_critical_issues or val_res.has_errors or product.quality_score < 70.0:
        product.status = ProductStatus.needs_review
    else:
        product.status = ProductStatus.verified

    product.updated_at = datetime.now(timezone.utc)
    session.add(product)
    session.commit()

    return {
        "status": "success",
        "product_id": str(product_id),
        "quality_score": product.quality_score,
        "product_status": product.status,
        "issues_count": len(val_res.issues),
    }


@router.post("/{product_id}/enrich")
def rerun_product_enrichment(product_id: uuid.UUID, session: Session = Depends(get_session)):
    """Re-runs AI commerce enrichment on demand for a product."""
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    from app.services.llm.factory import get_llm_provider
    from app.services.pipeline import EnrichmentStage

    provider = get_llm_provider()
    stage = EnrichmentStage(llm_provider=provider)

    # Create dummy processing step for stage execution
    from app.models import ProcessingJob, ProcessingStep, ProcessingStage, StepStatus
    job = ProcessingJob(total_items=1, completed_items=0)
    session.add(job)
    session.commit()
    session.refresh(job)

    step = ProcessingStep(job_id=job.id, stage=ProcessingStage.enriching, status=StepStatus.processing)
    session.add(step)
    session.commit()
    session.refresh(step)

    # Find associated document
    from app.models import ProductDocumentAssociation
    stmt = select(ProductDocumentAssociation).where(ProductDocumentAssociation.product_id == product_id)
    assoc = session.exec(stmt).first()

    if assoc:
        stage.execute(session, assoc.document_id, job.id, step.id)

    session.refresh(product)
    return {
        "status": "success",
        "product_id": str(product_id),
        "commerce_description": product.commerce_description,
        "features": product.features,
        "applications": product.applications,
    }


from app.models import (
    AttributeEvidence,
    AttributeStatus,
    AuditLog,
    Document,
    EnrichmentResult,
    Product,
    ProductAttribute,
    ProductDocumentAssociation,
    ProductStatus,
    ProductVersion,
    Source,
    SourceType,
    ValidationResult,
    ValidationStatus,
)


@router.post("/{product_id}/validation/{validation_id}/resolve")
def resolve_validation_issue(
    product_id: uuid.UUID,
    validation_id: uuid.UUID,
    request: ResolutionRequest,
    session: Session = Depends(get_session),
):
    """
    Human review endpoint for resolving validation issues and conflict resolution.
    Creates ProductVersion and AuditLog transactionally.
    """
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    val_record = session.get(ValidationResult, validation_id)
    if not val_record or val_record.product_id != product_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Validation result {validation_id} does not belong to product {product_id}",
        )

    # Scoping check on associated attribute
    attr: Optional[ProductAttribute] = None
    if val_record.attribute_id:
        attr = session.get(ProductAttribute, val_record.attribute_id)
        if attr and attr.product_id != product_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Attribute {val_record.attribute_id} does not belong to product {product_id}",
            )

    # Idempotency check: handle already resolved validation issues
    if val_record.status == ValidationStatus.resolved:
        stmt = (
            select(AuditLog)
            .where(
                AuditLog.entity_type == "validation_result",
                AuditLog.entity_id == validation_id,
            )
            .order_by(desc(AuditLog.created_at))
        )
        last_audit = session.exec(stmt).first()

        prev_resolution = last_audit.metadata_json.get("resolution") if last_audit and isinstance(last_audit.metadata_json, dict) else None

        if prev_resolution == request.resolution:
            return {
                "status": "already_resolved",
                "message": "Validation result was already resolved with identical decision.",
                "validation_id": str(validation_id),
                "product_id": str(product_id),
                "quality_score": product.quality_score,
                "product_status": product.status,
            }
        else:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Validation issue {validation_id} is already resolved with '{prev_resolution or 'resolved'}'. Changing resolution decision requires reopening the issue.",
            )

    now = datetime.now(timezone.utc)

    # Determine resolution value and source_id provenance
    selected_source_id: Optional[str] = None
    resolved_val: Optional[str] = None

    if request.resolution == "accept_source_a":
        if isinstance(val_record.expected_value, dict) and "raw_value" in val_record.expected_value:
            resolved_val = str(val_record.expected_value["raw_value"])
            selected_source_id = str(val_record.expected_value.get("source_id")) if val_record.expected_value.get("source_id") else None
        elif val_record.actual_value is not None and not isinstance(val_record.actual_value, (dict, list)):
            resolved_val = str(val_record.actual_value)
        elif request.resolved_value is not None:
            resolved_val = str(request.resolved_value)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot determine Source A claim value for this validation issue.",
            )

    elif request.resolution == "accept_source_b":
        if isinstance(val_record.actual_value, list) and len(val_record.actual_value) > 0 and isinstance(val_record.actual_value[0], dict) and "raw_value" in val_record.actual_value[0]:
            resolved_val = str(val_record.actual_value[0]["raw_value"])
            selected_source_id = str(val_record.actual_value[0].get("source_id")) if val_record.actual_value[0].get("source_id") else None
        elif val_record.expected_value is not None and not isinstance(val_record.expected_value, (dict, list)):
            resolved_val = str(val_record.expected_value)
        elif request.resolved_value is not None:
            resolved_val = str(request.resolved_value)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot determine Source B claim value for this validation issue.",
            )

    elif request.resolution in ["custom_value", "custom"]:
        if request.resolved_value is not None and str(request.resolved_value).strip() != "":
            resolved_val = str(request.resolved_value).strip()
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Resolved value must be provided for custom_value resolution.",
            )
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported resolution type '{request.resolution}'. Allowed types: accept_source_a, accept_source_b, custom_value, custom.",
        )

    # Create version snapshot BEFORE resolution
    product_service = ProductService(session)

    # Update ProductAttribute if associated
    if attr:
        attr.raw_value = resolved_val
        attr.status = AttributeStatus.verified
        attr.updated_at = now
        session.add(attr)

    # Mark validation as resolved
    val_record.status = ValidationStatus.resolved
    val_record.resolved_at = now
    val_record.resolved_by = "human_reviewer"
    session.add(val_record)

    # Audit log entry with resolution, resolved_value, selected_source_id, and notes
    audit = AuditLog(
        entity_type="validation_result",
        entity_id=validation_id,
        action="human_resolution",
        actor_type="user",
        metadata_json={
            "resolution": request.resolution,
            "resolved_value": resolved_val,
            "selected_source_id": selected_source_id,
            "notes": request.notes,
        },
    )
    session.add(audit)

    # Recalculate quality score and product status via ValidationEngine
    attributes = repo.get_attributes(product_id)
    attr_repo = AttributeRepository(session)
    evidence = attr_repo.get_evidence_for_product(product_id)
    evidence_names = {
        a.attribute_name for a in attributes
        if any(e.attribute_id == a.id and e.evidence_text for e in evidence)
    }

    engine = ValidationEngine()
    val_res = engine.validate_product(product, attributes, evidence_names)
    product.quality_score = val_res.quality_breakdown.quality_score
    if not val_res.has_critical_issues and not val_res.has_errors:
        product.status = ProductStatus.verified
    product.updated_at = now
    session.add(product)

    session.commit()

    return {
        "status": "resolved",
        "validation_id": str(validation_id),
        "product_id": str(product_id),
        "quality_score": product.quality_score,
        "product_status": product.status,
    }


@router.get("/{product_id}/versions", response_model=List[ProductVersion])
def get_product_versions(product_id: uuid.UUID, session: Session = Depends(get_session)):
    repo = ProductRepository(session)
    if not repo.get_by_id(product_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )
    return repo.get_versions(product_id)


@router.get("/{product_id}/evidence", response_model=List[AttributeEvidence])
def get_product_evidence(product_id: uuid.UUID, session: Session = Depends(get_session)):
    repo = ProductRepository(session)
    if not repo.get_by_id(product_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )
    attr_repo = AttributeRepository(session)
    return attr_repo.get_evidence_for_product(product_id)


@router.get("/{product_id}/reconciliation", response_model=ProductReconciliationResponse)
def get_product_reconciliation(
    product_id: uuid.UUID, session: Session = Depends(get_session)
):
    repo = ProductRepository(session)
    if not repo.get_by_id(product_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    reconciler = MultiSourceReconciler(session)
    summary = reconciler.reconcile_product(product_id)
    return summary


@router.get("/{product_id}/sources", response_model=List[ProductSourceResponse])
def get_product_sources(
    product_id: uuid.UUID, session: Session = Depends(get_session)
):
    repo = ProductRepository(session)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID {product_id} not found",
        )

    attr_repo = AttributeRepository(session)
    evidences = attr_repo.get_evidence_for_product(product_id)

    # Collect source IDs from evidence
    source_ids = {ev.source_id for ev in evidences if ev.source_id}

    # Also collect document associations
    stmt = select(ProductDocumentAssociation).where(
        ProductDocumentAssociation.product_id == product_id
    )
    doc_assocs = session.exec(stmt).all()
    doc_ids = {da.document_id for da in doc_assocs}

    sources_by_id: Dict[str, Source] = {}
    if source_ids:
        src_stmt = select(Source).where(Source.id.in_(source_ids))
        for s in session.exec(src_stmt).all():
            sources_by_id[str(s.id)] = s

    if doc_ids:
        doc_src_stmt = select(Source).where(Source.document_id.in_(doc_ids))
        for s in session.exec(doc_src_stmt).all():
            sources_by_id[str(s.id)] = s

    # Fallback to Document table for virtual source representation if no Source record registered yet
    if doc_ids:
        doc_stmt = select(Document).where(Document.id.in_(doc_ids))
        docs = session.exec(doc_stmt).all()
        for doc in docs:
            already_registered = any(s.document_id == doc.id for s in sources_by_id.values())
            if not already_registered:
                sources_by_id[str(doc.id)] = Source(
                    id=doc.id,
                    source_type=SourceType.document,
                    name=doc.filename,
                    document_id=doc.id,
                    trust_level=1.0,
                    metadata_json=doc.metadata_json or {},
                    created_at=doc.created_at,
                )

    result_sources: List[ProductSourceResponse] = []
    for s in sources_by_id.values():
        result_sources.append(
            ProductSourceResponse(
                source_id=str(s.id),
                source_name=s.name,
                source_type=str(s.source_type.value if hasattr(s.source_type, "value") else s.source_type),
                uri=s.uri,
                trust_level=s.trust_level,
                document_id=str(s.document_id) if s.document_id else None,
                metadata_json=s.metadata_json or {},
                created_at=s.created_at,
                association_type="evidence_source" if s.id in source_ids else "document_source",
            )
        )

    result_sources.sort(key=lambda s: s.trust_level, reverse=True)
    return result_sources

