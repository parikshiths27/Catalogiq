import io
import csv
import openpyxl
from fastapi.testclient import TestClient
from sqlmodel import Session

from app.main import app
from app.db.session import engine
from app.models import (
    Product, ProductStatus, ProductAttribute, AttributeDataType,
    AttributeStatus, ValidationResult, ValidationType, ValidationSeverity,
    ValidationStatus, Document, DocumentStatus
)

def test_export_endpoints():
    client = TestClient(app)

    # 1. Test CSV export when empty
    res_csv = client.get("/api/v1/products/export?format=csv")
    assert res_csv.status_code == 200
    assert "text/csv" in res_csv.headers.get("content-type", "")
    assert "attachment; filename=CatalogIQ_Export.csv" in res_csv.headers.get("content-disposition", "")

    # 2. Test Excel export when empty
    res_xlsx = client.get("/api/v1/products/export?format=xlsx")
    assert res_xlsx.status_code == 200
    assert "spreadsheetml" in res_xlsx.headers.get("content-type", "")

    # 3. Create a test product
    with Session(engine) as session:
        prod = Product(
            sku="TEST-SKU-999",
            brand="Siemens Test",
            product_name="Siemens Industrial Test Motor",
            category="Electric Motor",
            status=ProductStatus.needs_review,
            quality_score=65.0,
            description="Test engineering description",
            commerce_description="High-performance industrial 3-phase induction motor",
            features=["IP55 Protection", "Cast Iron Frame"],
            applications=["Pumps", "Conveyors"],
        )
        session.add(prod)
        session.flush()

        attr = ProductAttribute(
            product_id=prod.id,
            attribute_name="power_output",
            display_name="Rated Power",
            raw_value="15",
            normalized_value=15.0,
            unit="kW",
            data_type=AttributeDataType.numeric,
            confidence=0.95,
            status=AttributeStatus.verified,
            source_type="document",
        )
        session.add(attr)

        val = ValidationResult(
            product_id=prod.id,
            validation_type=ValidationType.low_confidence,
            severity=ValidationSeverity.warning,
            status=ValidationStatus.open,
            message="Confidence score below threshold for cooling_type",
        )
        session.add(val)
        session.commit()
        prod_id = prod.id

    try:
        # Test CSV export with product
        res_csv = client.get("/api/v1/products/export?format=csv")
        assert res_csv.status_code == 200
        content = res_csv.content.decode("utf-8")
        assert "TEST-SKU-999" in content
        assert "Siemens Test" in content
        assert "Rated Power" in content
        assert "15.0 kW" in content or "15 kW" in content

        # Test Excel export with product
        res_xlsx = client.get("/api/v1/products/export?format=xlsx")
        assert res_xlsx.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(res_xlsx.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "SKU" in headers
        assert "Brand" in headers
        assert "Rated Power" in headers
        rows = list(ws.iter_rows(values_only=True))
        assert any("TEST-SKU-999" in str(row) for row in rows)

    finally:
        # Clean up test product
        with Session(engine) as session:
            for m, col in [(ValidationResult, ValidationResult.product_id), (ProductAttribute, ProductAttribute.product_id), (Product, Product.id)]:
                items = session.query(m).filter(col == prod_id).all()
                for it in items:
                    session.delete(it)
            session.commit()

    print("All export endpoint tests PASSED successfully!")


def test_clear_all_documents():
    client = TestClient(app)
    res = client.delete("/api/v1/documents/clear-all")
    assert res.status_code == 200
    data = res.json()
    assert data["status"] == "cleared"
    print("Clear all documents endpoint test PASSED successfully!")


if __name__ == "__main__":
    test_export_endpoints()
    test_clear_all_documents()
